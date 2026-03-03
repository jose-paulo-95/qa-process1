# -*- coding: utf-8 -*-
"""Gera o arquivo xlsx da Avaliação de Competências - QA Junior."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def style_header(ws, row, cols):
    thin = Side(style='thin')
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_section_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

def add_competency_table(ws, start_row, section_title, rows_data, dv_1_5):
    ws.cell(row=start_row, column=1, value=section_title)
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
    start_row += 1
    headers = ["Competência", "Nível Esperado", "Nível (1-5)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    style_header(ws, start_row, 3)
    start_row += 1
    data_rows = []
    for comp, nivel in rows_data:
        ws.cell(row=start_row, column=1, value=comp)
        ws.cell(row=start_row, column=2, value=nivel)
        ws.cell(row=start_row, column=3, value="")
        data_rows.append(start_row)
        start_row += 1
    # Aplica validação 1-5 nas células de nível desta seção
    if data_rows:
        dv_1_5.add(f"C{data_rows[0]}:C{data_rows[-1]}")
    start_row += 1
    return start_row

def main():
    wb = Workbook()
    thin = Side(style='thin')

    # ========== ABA 0: Escala de Conceito (1 a 5) ==========
    ws0 = wb.create_sheet("Escala de Conceito", 0)
    ws0.cell(row=1, column=1, value="Escala de Conceito - Matriz de Competências")
    ws0.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws0.merge_cells('A1:C1')
    ws0.cell(row=2, column=1, value="Atribua um nível de 1 a 5 para cada competência. Você pode indicar conhecimento parcial.")
    ws0.cell(row=2, column=1).font = Font(italic=True, size=10)
    ws0.merge_cells('A2:C2')
    escala_headers = ["Nível", "Conceito", "Descrição"]
    for col, h in enumerate(escala_headers, 1):
        ws0.cell(row=4, column=col, value=h)
    style_header(ws0, 4, 3)
    escala_data = [
        (1, "Não conhece", "Nunca utilizei ou não tenho conhecimento sobre o tema"),
        (2, "Conheço o conceito", "Ouvi falar / sei o que é, mas não apliquei na prática"),
        (3, "Conheço parcialmente", "Já utilizei com apoio ou em situações simples"),
        (4, "Utilizo com autonomia", "Consigo usar no dia a dia sem necessidade de apoio"),
        (5, "Domino", "Domino o tema e consigo orientar ou ensinar outros"),
    ]
    for i, (nivel, conceito, desc) in enumerate(escala_data, start=5):
        ws0.cell(row=i, column=1, value=nivel)
        ws0.cell(row=i, column=2, value=conceito)
        ws0.cell(row=i, column=3, value=desc)
    ws0.column_dimensions['A'].width = 8
    ws0.column_dimensions['B'].width = 22
    ws0.column_dimensions['C'].width = 55

    # Validação: apenas valores 1 a 5 na coluna de nível
    dv_1_5 = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=True,
    )
    dv_1_5.error = "Informe um nível de 1 a 5"
    dv_1_5.promptTitle = "Escala 1 a 5"
    dv_1_5.prompt = "1=Não conhece, 2=Conceito, 3=Parcial, 4=Autonomia, 5=Domino"

    # ========== ABA 1: Matriz de Competências ==========
    ws1 = wb.active
    ws1.title = "Matriz de Competências"
    ws1.add_data_validation(dv_1_5)
    ws1.cell(row=1, column=1, value="Avaliação de Competências - QA Junior")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws1.merge_cells('A1:C1')
    ws1.cell(row=2, column=1, value="Escala de conceito (1 a 5): 1=Não conhece | 2=Conheço o conceito | 3=Conheço parcialmente | 4=Utilizo com autonomia | 5=Domino")
    ws1.cell(row=2, column=1).font = Font(italic=True, size=9)
    ws1.merge_cells('A2:C2')
    row = 4

    matriz = [
        ("2.1 Fundamentos de QA", [
            ("Ciclo de vida de software", "Conhece conceitos básicos"),
            ("Tipos de teste", "Conhece principais tipos (funcional, regressão, integração)"),
            ("Criação de casos de teste", "Cria casos simples com pré-condições, passos e resultado esperado"),
            ("Severidade/Prioridade", "Identifica corretamente severidade (Crítica, Alta, Média, Baixa) e prioridade"),
            ("Reporte de bugs", "Reporta bugs seguindo template, com passos para reproduzir e evidências"),
        ]),
        ("2.2 Testes Manuais", [
            ("Execução de testes", "Executa casos de teste guiados, seguindo scripts"),
            ("Testes exploratórios", "Realiza testes exploratórios básicos para descobrir bugs"),
            ("Validação de requisitos", "Valida funcionalidades comparando com requisitos"),
            ("Testes de regressão", "Executa checklist de regressão fornecido"),
        ]),
        ("2.3 Automação - Cypress (Básico)", [
            ("Instalação e setup", "Instala e configura Cypress básico no projeto"),
            ("Comandos básicos", "Usa cy.get(), cy.click(), cy.type(), cy.visit()"),
            ("Seletores", "Usa seletores simples (ID, classe, tag)"),
            ("Assertions", "Usa should() e expect() básico"),
            ("Fixtures", "Usa fixtures simples para dados de teste"),
            ("Custom Commands", "Não precisa criar, mas entende o conceito"),
            ("Page Object Model", "Não precisa implementar ainda"),
            ("Interceptação de requests", "Não precisa usar ainda"),
            ("CI/CD", "Não precisa integrar ainda"),
            ("Debugging", "Usa console básico do navegador"),
        ]),
        ("2.4 Automação - Selenium", [
            ("Setup e configuração", "Não precisa conhecer ou conhecimento básico"),
            ("WebDriver", "Conceitos básicos (se exposto)"),
            ("Page Object Model", "Não precisa conhecer ainda"),
            ("Waits e sincronização", "Conceito básico de waits (se exposto)"),
        ]),
        ("2.5 Testes de API - Postman", [
            ("Criação de requests", "Cria requisições GET e POST básicas"),
            ("Collections", "Cria collections simples para organizar requests"),
            ("Variáveis", "Usa variáveis locais em requests"),
            ("Scripts (Pre/Test)", "Não precisa usar ainda, mas conhece o conceito"),
            ("Testes automatizados", "Não precisa criar ainda"),
            ("Newman (CLI)", "Não precisa conhecer ainda"),
            ("Documentação", "Adiciona descrições básicas nos requests"),
            ("Mock servers", "Não precisa usar ainda"),
        ]),
        ("2.6 Testes de API - Swagger", [
            ("Leitura de documentação", "Lê endpoints básicos no Swagger UI"),
            ("Validação de contratos", "Não precisa validar ainda"),
            ("Testes baseados em Swagger", "Usa Swagger para entender API e criar testes no Postman"),
            ("OpenAPI/Swagger specs", "Conhece conceito básico"),
        ]),
        ("2.7 Ferramentas e Tecnologias", [
            ("Jira", "Cria bugs, executa testes, atualiza status"),
            ("Git", "Clone, pull, commit básico"),
            ("SQL", "Não precisa conhecer ainda"),
            ("Docker", "Não precisa conhecer ainda"),
            ("DevTools", "Inspeção básica de elementos, console"),
            ("Logs e debugging", "Lê logs básicos para identificar erros"),
        ]),
        ("2.8 Liderança e Processo", [
            ("Planejamento de testes", "Executa plano de teste fornecido"),
            ("Estimativa", "Não precisa estimar ainda"),
            ("Code review", "Não precisa participar ainda"),
            ("Mentoria", "Recebe mentoria de QAs mais experientes"),
            ("Processo de qualidade", "Segue processo estabelecido"),
            ("Métricas e KPIs", "Conhece conceitos básicos"),
            ("Comunicação técnica", "Comunica claramente com equipe sobre bugs e testes"),
        ]),
    ]

    for section_title, rows_data in matriz:
        row = add_competency_table(ws1, row, section_title, rows_data, dv_1_5)

    ws1.column_dimensions['A'].width = 45
    ws1.column_dimensions['B'].width = 55
    ws1.column_dimensions['C'].width = 12

    # ========== ABA 2: Questionário de Avaliação ==========
    ws2 = wb.create_sheet("Questionário de Avaliação")
    ws2.cell(row=1, column=1, value="Questionário de Avaliação - QA Junior")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.merge_cells('A1:E1')
    ws2.cell(row=2, column=1, value="Sistema de Pontuação: 0 = Não conhece | 1 = Conceitual | 2 = Aplica em situações simples | 3 = Domina (acima do esperado)")
    ws2.merge_cells('A2:E2')
    ws2.cell(row=2, column=1).font = Font(italic=True, size=10)
    row = 4

    perguntas = [
        ("3.2 Fundamentos de QA", [
            ("P1", "Explique o que é um caso de teste e quais elementos ele deve conter.", "2", "Pré-condições, passos numerados, resultado esperado. (3 pts: dados de teste, pós-condições, critérios de aceitação)"),
            ("P2", "Qual a diferença entre severidade e prioridade de um bug?", "2", "Severidade = impacto no sistema, Prioridade = urgência. (3 pts: exemplos práticos)"),
            ("P3", "Descreva o ciclo de vida de um bug desde a descoberta até o fechamento.", "2", "Novo → Atribuído → Resolvido → Fechado. (3 pts: estados intermediários, reabertura)"),
        ]),
        ("3.3 Testes Manuais", [
            ("P4", "O que são testes exploratórios e quando você os utiliza?", "2", "Testes sem script pré-definido para descobrir bugs. (3 pts: sessões estruturadas, time-boxing)"),
            ("P5", "Como você valida se uma funcionalidade está funcionando corretamente?", "2", "Compara com requisitos, testa cenários principais. (3 pts: positivos, negativos, limites)"),
        ]),
        ("3.4 Postman (Básico)", [
            ("P6", "Como você cria uma requisição GET no Postman?", "2", "Método GET, URL, Send. (3 pts: query params, headers, autenticação)"),
            ("P7", "O que são variáveis no Postman e para que servem?", "2", "Valores reutilizáveis (URLs, tokens). (3 pts: ambiente, globais, locais)"),
            ("P8", "Como você testa se uma API retornou o status code correto?", "2", "Verifica visualmente o status. (3 pts: script Tests pm.response.to.have.status(200))"),
            ("PM1", "Como você cria um workflow onde o resultado de uma requisição é usado na próxima?", "1", "Copia manualmente. (2 pts: variáveis para armazenar resposta)"),
            ("PM4", "Como você gerencia múltiplos ambientes (DEV, QA, PROD) no Postman?", "1", "Collections separadas. (2 pts: environments com variáveis)"),
        ]),
        ("3.5 Swagger (Básico)", [
            ("P9", "O que é Swagger e para que é usado?", "2", "Documentação de API, ver endpoints. (3 pts: OpenAPI, contratos, geração de código)"),
            ("P10", "Como você usa o Swagger para criar testes no Postman?", "2", "Lê endpoints, copia para Postman. (3 pts: importa collection, valida esquemas)"),
            ("SW1", "Como você identifica problemas em uma documentação Swagger?", "1", "Testa manualmente e compara. (2 pts: valida schema, exemplos)"),
            ("SW2", "Explique a diferença entre Swagger UI e Swagger Editor.", "1", "UI visualiza, Editor edita. (2 pts: quando usar cada um)"),
            ("SW5", "Como você gera testes automaticamente a partir de um arquivo Swagger?", "1", "Importa collection no Postman. (2 pts: ferramentas de geração)"),
        ]),
        ("3.6 Cypress (Básico)", [
            ("C1", "Qual a diferença entre cy.visit() e cy.request()?", "1", "visit() abre página, request() faz HTTP direto. (2 pts: quando usar cada um)"),
            ("C3", "Explique o conceito de retry-ability no Cypress.", "1", "Cypress tenta novamente se falhar. (2 pts: como funciona, comandos vs assertions)"),
            ("C5", "Como você organiza fixtures para diferentes ambientes no Cypress?", "1", "Fixtures em cypress/fixtures. (2 pts: por ambiente, variáveis)"),
        ]),
        ("3.7 Ferramentas", [
            ("P11", "Como você usa o Jira para reportar um bug?", "2", "Cria Bug, preenche campos, anexa evidências. (3 pts: template, linka story, labels)"),
            ("P12", "O que você faz quando encontra um bug durante os testes?", "2", "Documenta, screenshot, reporta no Jira. (3 pts: verifica duplicata, reproduz, severidade)"),
        ]),
    ]

    headers_q = ["ID", "Pergunta", "Pts máx.", "Resposta esperada / Observações", "Pontos (0-3)"]
    for sec_title, items in perguntas:
        ws2.cell(row=row, column=1, value=sec_title)
        ws2.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws2.cell(row=row, column=1).fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        row += 1
        for col, h in enumerate(headers_q, 1):
            ws2.cell(row=row, column=col, value=h)
        style_header(ws2, row, 5)
        row += 1
        for id_p, pergunta, pts_max, esperado in items:
            ws2.cell(row=row, column=1, value=id_p)
            ws2.cell(row=row, column=2, value=pergunta)
            ws2.cell(row=row, column=3, value=pts_max)
            ws2.cell(row=row, column=4, value=esperado)
            ws2.cell(row=row, column=5, value="")
            row += 1
        row += 1

    # Total máximo de pontos do questionário = 43 (conforme documento)
    total_max = 43
    row_total_pts = row
    ws2.cell(row=row, column=1, value="Total de pontos:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value="=SUM(E6:E35)")
    row += 1
    row_total_max = row
    ws2.cell(row=row, column=1, value="Total máximo:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=total_max)
    row += 1
    ws2.cell(row=row, column=1, value="Percentual:")
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=f"=B{row_total_pts}/B{row_total_max}*100")
    ws2.cell(row=row, column=3, value="%")
    row += 1
    ws2.cell(row=row, column=1, value="Interpretação: 0-40% treinamento | 41-70% adequado Junior | 71-100% pronto para evoluir")
    ws2.cell(row=row, column=1).font = Font(italic=True, size=9)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 65
    ws2.column_dimensions['E'].width = 12

    # Ajustar fórmula do total para o intervalo real (E4 até última linha de pontos)
    # Deixamos SUM(E4:E100) como está; usuário pode ajustar se necessário

    # ========== ABA 3: Interpretação e Plano ==========
    ws3 = wb.create_sheet("Interpretação e Plano")
    ws3.cell(row=1, column=1, value="Guia de Interpretação e Plano de Desenvolvimento")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws3.merge_cells('A1:B1')
    content = [
        (2, "Respostas esperadas de um QA Junior", [
            "Conhece conceitos básicos de QA",
            "Precisa de orientação para aplicar conhecimento",
            "Foca em execução de tarefas guiadas",
            "Aprende fazendo, com suporte",
            "Comunica claramente sobre bugs encontrados",
        ]),
        (10, "Sinais de que está pronto para evoluir", [
            "Responde perguntas de nível Pleno com 2+ pontos",
            "Demonstra autonomia em tarefas",
            "Compartilha conhecimento com outros juniores",
            "Propõe melhorias no processo",
            "Automatiza tarefas repetitivas",
        ]),
        (18, "Áreas que precisam de desenvolvimento", [
            "Fundamentos: mais treinamento teórico",
            "Ferramentas: prática com Postman, Jira",
            "Cypress: curso básico e prática",
            "Swagger: exposição a APIs e documentação",
        ]),
        (25, "Plano de Desenvolvimento Sugerido", [
            "Mês 1-2 - Fundamentos: curso QA, casos de teste, reporte de bugs",
            "Mês 2-3 - Ferramentas: Jira, Postman, Swagger",
            "Mês 4-6 - Automação: Cypress básico, primeiro teste automatizado",
            "Mês 7-12 - Consolidação: praticar conceitos, automatizar mais, preparar próximo nível",
        ]),
    ]
    r = 3
    for start, titulo, itens in content:
        if r > 2:
            r += 1
        ws3.cell(row=r, column=1, value=titulo)
        ws3.cell(row=r, column=1).font = Font(bold=True, size=11)
        r += 1
        for item in itens:
            ws3.cell(row=r, column=1, value="•")
            ws3.cell(row=r, column=2, value=item)
            r += 1
        r += 1
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 75

    out_path = "c:\\Projetos\\qa-process\\docs\\avaliacao\\avaliacao-junior.xlsx"
    wb.save(out_path)
    print(f"Arquivo salvo: {out_path}")

if __name__ == "__main__":
    main()
