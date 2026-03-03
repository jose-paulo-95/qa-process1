# -*- coding: utf-8 -*-
"""Gera o formulário XLSX de Autoavaliação para envio ao QA Junior."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

DIR = Path(__file__).parent
OUT_PATH = DIR / "formulario-autoavaliacao-qa-junior.xlsx"


def style_header(ws, row, cols):
    thin = Side(style="thin")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_section(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")


def main():
    wb = Workbook()
    thin = Side(style="thin")

    # Validação 1-5
    dv_1_5 = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=True,
    )
    dv_1_5.error = "Informe um nível de 1 a 5"
    dv_1_5.promptTitle = "Escala 1 a 5"
    dv_1_5.prompt = "1=Não conhece, 2=Conceito, 3=Parcial, 4=Autonomia, 5=Domino"

    # Aba oculta com opções para lista suspensa
    ws_opcoes = wb.create_sheet("_Opcoes", 3)
    ws_opcoes.sheet_state = "hidden"
    for i, opc in enumerate(["0-6 meses", "6 meses - 1 ano", "1-2 anos"], 1):
        ws_opcoes.cell(row=i, column=1, value=opc)
    dv_exp = DataValidation(type="list", formula1=f"'_Opcoes'!$A$1:$A$3", allow_blank=True)
    dv_exp.error = "Selecione uma opção da lista"
    dv_exp.promptTitle = "Tempo de experiência"

    # ========== ABA 1: Instruções e Escala ==========
    ws0 = wb.active
    ws0.title = "Instruções"
    ws0.cell(row=1, column=1, value="Formulário de Autoavaliação - QA Junior")
    ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws0.merge_cells("A1:C1")
    ws0.cell(row=2, column=1, value="Preencha este formulário antes da sua reunião de avaliação. Suas respostas ajudarão a personalizar seu Plano de Desenvolvimento Individual (PDI).")
    ws0.cell(row=2, column=1).font = Font(italic=True, size=10)
    ws0.merge_cells("A2:C2")
    ws0.cell(row=4, column=1, value="Escala de Conceito (1 a 5) - Use na Matriz de Competências:")
    ws0.cell(row=4, column=1).font = Font(bold=True)
    escala = [
        (1, "Não conhece", "Nunca utilizei ou não tenho conhecimento sobre o tema"),
        (2, "Conheço o conceito", "Ouvi falar / sei o que é, mas não apliquei na prática"),
        (3, "Conheço parcialmente", "Já utilizei com apoio ou em situações simples"),
        (4, "Utilizo com autonomia", "Consigo usar no dia a dia sem necessidade de apoio"),
        (5, "Domino", "Domino o tema e consigo orientar ou ensinar outros"),
    ]
    for i, (n, c, d) in enumerate(escala, start=5):
        ws0.cell(row=i, column=1, value=n)
        ws0.cell(row=i, column=2, value=c)
        ws0.cell(row=i, column=3, value=d)
    ws0.column_dimensions["A"].width = 8
    ws0.column_dimensions["B"].width = 22
    ws0.column_dimensions["C"].width = 55

    # ========== ABA 2: Perfil e Auto-Reflexão ==========
    ws1 = wb.create_sheet("Perfil e Auto-Reflexão", 1)
    ws1.add_data_validation(dv_exp)

    row = 1
    ws1.cell(row=row, column=1, value="PERFIL E CONTEXTO")
    ws1.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    perfil = [
        ("Nome completo:", "B", 40),
        ("Data desta autoavaliação:", "B", 15),
        ("Tempo de experiência em QA:", "B", 20),
        ("Projetos ou áreas em que você atuou:", "B", 50),
        ("Nome do seu líder/avaliador:", "B", 30),
    ]
    for label, col_resp, width in perfil:
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=1).font = Font(bold=True)
        resp_cell = ws1.cell(row=row, column=2, value="")
        resp_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        resp_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if "Data" in label:
            resp_cell.number_format = "DD/MM/AAAA"
        elif "Tempo" in label:
            dv_exp.add(f"B{row}")
        ws1.column_dimensions["B"].width = width
        row += 1

    row += 2
    ws1.cell(row=row, column=1, value="AUTO-REFLEXÃO")
    ws1.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws1.cell(row=row + 1, column=1, value="As perguntas abaixo ajudam a entender suas metas e preferências. Não há resposta certa ou errada - seja sincero.")
    ws1.cell(row=row + 1, column=1).font = Font(italic=True, size=9)
    ws1.merge_cells(f"A{row+1}:B{row+1}")
    row += 3

    reflexao = [
        ("Quais são suas metas de carreira para os próximos 12 meses? O que você quer alcançar?", 6),
        ("Quais áreas de QA mais te interessam? (liste até 3)", 4),
        ("Quais são seus principais desafios atuais no dia a dia? O que mais te dificulta?", 6),
        ("Como você prefere aprender? (Curso online, mentoria, prática hands-on, leitura, etc.)", 4),
        ("O que você gostaria de desenvolver primeiro? Priorize até 3 itens.", 6),
    ]
    for label, height in reflexao:
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=1).font = Font(bold=True)
        ws1.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        resp = ws1.cell(row=row, column=2, value="")
        resp.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        resp.alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[row].height = height * 15
        row += 1

    ws1.column_dimensions["A"].width = 50
    ws1.column_dimensions["B"].width = 55

    # ========== ABA 3: Matriz de Competências ==========
    ws2 = wb.create_sheet("Matriz de Competências", 2)
    ws2.add_data_validation(dv_1_5)

    ws2.cell(row=1, column=1, value="MATRIZ DE COMPETÊNCIAS - AUTOAVALIAÇÃO")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")
    ws2.cell(row=2, column=1, value="Atribua um nível de 1 a 5 para cada competência. Na coluna 'Detalhes', você pode opcionalmente descrever seu conhecimento (ex.: ferramentas usadas, projetos, observações).")
    ws2.cell(row=2, column=1).font = Font(italic=True, size=9)
    ws2.merge_cells("A2:D2")
    row = 4

    # Cabeçalho da tabela
    headers = ["Competência", "Nível Esperado", "Nível (1-5)", "Detalhes (opcional)"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row += 1

    matriz = [
        ("2.1 Fundamentos de QA", [
            ("Ciclo de vida de software", "Conhece conceitos básicos"),
            ("Tipos de teste", "Conhece principais tipos (funcional, regressão, integração)"),
            ("Criação de casos de teste", "Cria casos simples com pré-condições, passos e resultado esperado"),
            ("Severidade/Prioridade", "Identifica corretamente severidade e prioridade"),
            ("Reporte de bugs", "Reporta bugs seguindo template, com passos para reproduzir e evidências"),
        ]),
        ("2.2 Testes Manuais", [
            ("Execução de testes", "Executa casos de teste guiados, seguindo scripts"),
            ("Testes exploratórios", "Realiza testes exploratórios básicos para descobrir bugs"),
            ("Validação de requisitos", "Valida funcionalidades comparando com requisitos"),
            ("Testes de regressão", "Executa checklist de regressão fornecido"),
        ]),
        ("2.3 Automação - Cypress (Básico)", [
            ("Instalação e setup", "Instala e configura Cypress básico no projeto"),
            ("Comandos básicos", "Usa cy.get(), cy.click(), cy.type(), cy.visit()"),
            ("Seletores", "Usa seletores simples (ID, classe, tag)"),
            ("Assertions", "Usa should() e expect() básico"),
            ("Fixtures", "Usa fixtures simples para dados de teste"),
            ("Custom Commands", "Não precisa criar, mas entende o conceito"),
            ("Page Object Model", "Não precisa implementar ainda"),
            ("Interceptação de requests", "Não precisa usar ainda"),
            ("CI/CD", "Não precisa integrar ainda"),
            ("Debugging", "Usa console básico do navegador"),
        ]),
        ("2.4 Automação - Selenium", [
            ("Setup e configuração", "Não precisa conhecer ou conhecimento básico"),
            ("WebDriver", "Conceitos básicos (se exposto)"),
            ("Page Object Model", "Não precisa conhecer ainda"),
            ("Waits e sincronização", "Conceito básico de waits (se exposto)"),
        ]),
        ("2.5 Testes de API - Postman", [
            ("Criação de requests", "Cria requisições GET e POST básicas"),
            ("Collections", "Cria collections simples para organizar requests"),
            ("Variáveis", "Usa variáveis locais em requests"),
            ("Scripts (Pre/Test)", "Não precisa usar ainda, mas conhece o conceito"),
            ("Testes automatizados", "Não precisa criar ainda"),
            ("Newman (CLI)", "Não precisa conhecer ainda"),
            ("Documentação", "Adiciona descrições básicas nos requests"),
            ("Mock servers", "Não precisa usar ainda"),
        ]),
        ("2.6 Testes de API - Swagger", [
            ("Leitura de documentação", "Lê endpoints básicos no Swagger UI"),
            ("Validação de contratos", "Não precisa validar ainda"),
            ("Testes baseados em Swagger", "Usa Swagger para entender API e criar testes no Postman"),
            ("OpenAPI/Swagger specs", "Conhece conceito básico"),
        ]),
        ("2.7 Ferramentas e Tecnologias", [
            ("Jira", "Cria bugs, executa testes, atualiza status"),
            ("Git", "Clone, pull, commit básico"),
            ("SQL", "Não precisa conhecer ainda"),
            ("Docker", "Não precisa conhecer ainda"),
            ("DevTools", "Inspeção básica de elementos, console"),
            ("Logs e debugging", "Lê logs básicos para identificar erros"),
        ]),
        ("2.8 Liderança e Processo", [
            ("Planejamento de testes", "Executa plano de teste fornecido"),
            ("Estimativa", "Não precisa estimar ainda"),
            ("Code review", "Não precisa participar ainda"),
            ("Mentoria", "Recebe mentoria de QAs mais experientes"),
            ("Processo de qualidade", "Segue processo estabelecido"),
            ("Métricas e KPIs", "Conhece conceitos básicos"),
            ("Comunicação técnica", "Comunica claramente com equipe sobre bugs e testes"),
        ]),
        ("2.9 Competências Comportamentais (Soft Skills)", [
            ("Comunicação clara (oral e escrita)", "Comunica bugs e resultados de forma objetiva"),
            ("Atenção aos detalhes", "Identifica inconsistências em evidências e documentação"),
            ("Trabalho em equipe", "Colabora com devs e outros QAs sem conflitos"),
            ("Pensamento crítico", "Questiona requisitos ambíguos, pensa em edge cases"),
            ("Curiosidade e aprendizado", "Busca aprender novas ferramentas e conceitos"),
            ("Empatia (advocacia do usuário)", "Considera a perspectiva do usuário ao testar"),
        ]),
    ]

    data_rows = []
    for section_title, rows_data in matriz:
        ws2.cell(row=row, column=1, value=section_title)
        ws2.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws2.cell(row=row, column=1).fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        for comp, nivel in rows_data:
            ws2.cell(row=row, column=1, value=comp)
            ws2.cell(row=row, column=2, value=nivel)
            cell_nivel = ws2.cell(row=row, column=3, value="")
            cell_nivel.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell_detalhes = ws2.cell(row=row, column=4, value="")
            cell_detalhes.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell_detalhes.alignment = Alignment(wrap_text=True, vertical="top")
            data_rows.append(row)
            row += 1
        row += 1

    if data_rows:
        dv_1_5.add(f"C{data_rows[0]}:C{data_rows[-1]}")

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 40

    wb.save(OUT_PATH)
    print(f"Formulário gerado: {OUT_PATH}")


if __name__ == "__main__":
    main()
